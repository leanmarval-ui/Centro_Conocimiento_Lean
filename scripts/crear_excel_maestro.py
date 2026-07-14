from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os

# Crear libro
wb = Workbook()

# Eliminar la hoja por defecto
wb.remove(wb.active)

# Nombre de las hojas y columnas
hojas = {
    "Herramientas": [
        "ID",
        "Nombre",
        "Categoría",
        "Descripción",
        "Responsable",
        "Estado",
        "Link",
        "Manual",
        "Observaciones"
    ],

    "Desarrollos": [
        "ID",
        "Nombre",
        "Objetivo",
        "Tecnología",
        "Responsable",
        "Estado",
        "GitHub",
        "Documentación",
        "Próximos pasos"
    ],

    "Indicadores": [
        "ID",
        "Nombre",
        "Plataforma",
        "Responsable",
        "Frecuencia",
        "Link",
        "Fuente de datos"
    ],

    "Procesos": [
        "ID",
        "Proceso",
        "Objetivo",
        "Responsable",
        "Entradas",
        "Salidas",
        "Documentación"
    ],

    "Documentación": [
        "ID",
        "Documento",
        "Tipo",
        "Responsable",
        "Ubicación",
        "Versión",
        "Fecha"
    ],

    "Equipo": [
        "Nombre",
        "Cargo",
        "Funciones",
        "Correo",
        "Teléfono",
        "Observaciones"
    ],

    "Roadmap": [
        "Proyecto",
        "Estado",
        "Prioridad",
        "Responsable",
        "Fecha estimada",
        "Comentarios"
    ]
}

# Estilo
color = PatternFill(
    start_color="1C4D78",
    end_color="1C4D78",
    fill_type="solid"
)

for nombre, columnas in hojas.items():

    ws = wb.create_sheet(title=nombre)

    for i, columna in enumerate(columnas, start=1):

        celda = ws.cell(row=1, column=i)

        celda.value = columna
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = color

        ws.column_dimensions[chr(64+i)].width = 25

# Crear carpeta data si no existe
os.makedirs("data", exist_ok=True)

# Guardar archivo
wb.save("data/Centro_Conocimiento_LEAN.xlsx")

print("✅ Archivo creado correctamente.")